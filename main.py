import flet as ft
import json
import urllib.request
import urllib.parse
import time
import os
import tempfile
from datetime import datetime

# Excel & PDF Libraries
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# ==================== SUPABASE CONFIG ====================
SUPABASE_URL = "https://nkllvhzebktydnqvjuoc.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im5rbGx2aHplYmt0eWRucXZqdW9jIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODYyODQzNDMsImV4cCI6MjEwMTg2MDM0M30.zIPGzDv5krWPUaIJzTP2BnKhSxU5LjJ0DraR46zFFLI"

# Direct HTTP Helper with Retry Logic (Fixes DNS & Network dropped packet errors)
def supabase_request(endpoint: str, method: str = "GET", data: dict = None, retries: int = 3):
    url = f"{SUPABASE_URL}/rest/v1/{endpoint}"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation"
    }
    
    body = json.dumps(data).encode("utf-8") if data is not None else None
    
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, data=body, headers=headers, method=method)
            with urllib.request.urlopen(req, timeout=10) as response:
                res_data = response.read().decode("utf-8")
                return json.loads(res_data) if res_data else []
        except urllib.error.HTTPError as e:
            err_msg = e.read().decode("utf-8")
            raise Exception(f"HTTP {e.code}: {err_msg}")
        except urllib.error.URLError as e:
            if attempt == retries - 1:
                raise Exception("Network Connection Error: Please check your internet connection.")
            time.sleep(1)
        except Exception as e:
            if attempt == retries - 1:
                raise Exception(str(e))
            time.sleep(1)

def main(page: ft.Page):
    page.title = "Kenooff Logistics"
    page.theme_mode = ft.ThemeMode.LIGHT
    page.horizontal_alignment = ft.CrossAxisAlignment.CENTER

    time.sleep(0.3)

    all_rows_cache = []
    active_filter_status = "All"

    def show_alert(message: str, bg_color):
        snack = ft.SnackBar(
            content=ft.Text(message, color=ft.Colors.WHITE),
            bgcolor=bg_color,
            open=True
        )
        page.overlay.append(snack)
        page.update()

    # ==================== FORM CONTROLS ====================
    vendor_input = ft.TextField(label="Vendor Name", border_radius=8)
    package_input = ft.TextField(label="Package Description", border_radius=8)
    qty_input = ft.TextField(label="Quantity", value="1", keyboard_type=ft.KeyboardType.NUMBER, border_radius=8)
    price_input = ft.TextField(label="Unit Price (₦)", value="0", keyboard_type=ft.KeyboardType.NUMBER, border_radius=8)
    
    location_input = ft.TextField(label="Delivery Location", border_radius=8)
    rider_input = ft.TextField(label="Assigned Rider", border_radius=8)
    delivery_charge = ft.TextField(label="Delivery Charges (₦)", value="0", keyboard_type=ft.KeyboardType.NUMBER, border_radius=8)
    
    payment_method = ft.Dropdown(
        label="Payment Method",
        options=[ft.dropdown.Option("Cash"), ft.dropdown.Option("Bank Transfer"), ft.dropdown.Option("POS")],
        value="Cash",
        border_radius=8
    )
    
    initial_status = ft.Dropdown(
        label="Initial Status",
        options=[ft.dropdown.Option("Pending"), ft.dropdown.Option("In Transit"), ft.dropdown.Option("Delivered")],
        value="Pending",
        border_radius=8
    )

    total_badge = ft.Text("Total Value: ₦0.00", size=18, weight=ft.FontWeight.BOLD, color=ft.Colors.GREEN_700)

    def run_calculation(e=None):
        try:
            qty = float(qty_input.value) if qty_input.value else 0
            price = float(price_input.value) if price_input.value else 0
            total = qty * price
            total_badge.value = f"Total Value: ₦{total:,.2f}"
            page.update()
            return total
        except Exception:
            total_badge.value = "Total Value: ₦0.00"
            page.update()
            return 0.0

    calc_btn = ft.ElevatedButton(
        content=ft.Row(
            [ft.Icon(ft.Icons.CALCULATE), ft.Text("Calculate Total Value")],
            alignment=ft.MainAxisAlignment.CENTER,
            tight=True,
        ),
        on_click=run_calculation,
        style=ft.ButtonStyle(color=ft.Colors.GREEN_700)
    )

    submit_btn_content = ft.Row(
        [ft.Text("Log & Save Package", color=ft.Colors.WHITE, weight=ft.FontWeight.BOLD)],
        alignment=ft.MainAxisAlignment.CENTER
    )
    
    submit_btn = ft.ElevatedButton(
        content=submit_btn_content,
        height=50,
        expand=True,
        style=ft.ButtonStyle(bgcolor=ft.Colors.BLUE_800)
    )

    def handle_submit(e):
        if not vendor_input.value or not package_input.value:
            show_alert("Please fill in Vendor Name and Package Description!", ft.Colors.RED_600)
            return

        submit_btn.disabled = True
        submit_btn_content.controls = [
            ft.ProgressRing(width=20, height=20, stroke_width=2, color=ft.Colors.WHITE),
            ft.Text(" Saving to Cloud...", color=ft.Colors.WHITE, weight=ft.FontWeight.BOLD)
        ]
        page.update()

        try:
            qty = float(qty_input.value) if qty_input.value else 0.0
            price = float(price_input.value) if price_input.value else 0.0
            del_charge = float(delivery_charge.value) if delivery_charge.value else 0.0
            total_val = qty * price

            data = {
                "vendor": vendor_input.value,
                "package_desc": package_input.value,
                "quantity": qty,
                "unit_price": price,
                "total_value": total_val,
                "location": location_input.value,
                "rider": rider_input.value,
                "delivery_charge": del_charge,
                "payment_method": payment_method.value,
                "status": initial_status.value
            }

            supabase_request("packages", method="POST", data=data)

            vendor_input.value = ""
            package_input.value = ""
            qty_input.value = "1"
            price_input.value = "0"
            location_input.value = ""
            rider_input.value = ""
            delivery_charge.value = "0"
            total_badge.value = "Total Value: ₦0.00"

            show_alert("Package Saved Online Successfully!", ft.Colors.GREEN_600)

        except Exception as ex:
            show_alert(f"Error saving to Cloud: {str(ex)}", ft.Colors.RED_600)

        finally:
            submit_btn.disabled = False
            submit_btn_content.controls = [
                ft.Text("Log & Save Package", color=ft.Colors.WHITE, weight=ft.FontWeight.BOLD)
            ]
            page.update()

    submit_btn.on_click = handle_submit

    form_view = ft.ListView(
        spacing=14,
        controls=[
            ft.Text("Kenooff Logistics Form", size=22, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_800),
            ft.Text("Log New Package Entry (Cloud Connected)", size=14, color=ft.Colors.GREY_600),
            ft.Divider(height=10),
            vendor_input,
            package_input,
            qty_input,
            price_input,
            ft.Container(height=5),
            total_badge,
            calc_btn,                  
            ft.Divider(height=10),
            location_input,
            rider_input,
            delivery_charge,
            payment_method,
            initial_status,
            ft.Container(height=10),
            ft.Row([submit_btn])
        ]
    )

    # ==================== ANALYTICS & FILTERING CONTROLS ====================
    total_rev_text = ft.Text("₦0.00", size=15, weight=ft.FontWeight.BOLD, color=ft.Colors.GREEN_800)
    total_count_text = ft.Text("0", size=15, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_800)
    delivered_count_text = ft.Text("0", size=15, weight=ft.FontWeight.BOLD, color=ft.Colors.GREEN_700)
    pending_count_text = ft.Text("0", size=15, weight=ft.FontWeight.BOLD, color=ft.Colors.ORANGE_700)

    analytics_row = ft.Row(
        controls=[
            ft.Container(
                content=ft.Column([ft.Text("Total Rev.", size=10, color=ft.Colors.GREY_600), total_rev_text], spacing=2),
                bgcolor=ft.Colors.GREEN_50, padding=8, border_radius=6, expand=True
            ),
            ft.Container(
                content=ft.Column([ft.Text("Total Pkgs", size=10, color=ft.Colors.GREY_600), total_count_text], spacing=2),
                bgcolor=ft.Colors.BLUE_50, padding=8, border_radius=6, expand=True
            ),
            ft.Container(
                content=ft.Column([ft.Text("Delivered", size=10, color=ft.Colors.GREY_600), delivered_count_text], spacing=2),
                bgcolor=ft.Colors.GREEN_100, padding=8, border_radius=6, expand=True
            ),
            ft.Container(
                content=ft.Column([ft.Text("Pending", size=10, color=ft.Colors.GREY_600), pending_count_text], spacing=2),
                bgcolor=ft.Colors.ORANGE_50, padding=8, border_radius=6, expand=True
            ),
        ],
        spacing=6
    )

    search_input = ft.TextField(
        hint_text="Search vendor, rider, location...", 
        prefix_icon=ft.Icons.SEARCH, 
        border_radius=8, 
        dense=True
    )

    # ==================== EXPORT FUNCTIONALITY ====================
    def generate_excel():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Logistics Report"

        headers = ["ID", "Vendor", "Package Description", "Qty", "Unit Price (₦)", "Total Value (₦)", "Location", "Rider", "Delivery Charge (₦)", "Payment", "Status", "Date"]
        ws.append(headers)

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r in all_rows_cache:
            ws.append([
                r.get("id"),
                r.get("vendor", ""),
                r.get("package_desc", ""),
                r.get("quantity", 0),
                r.get("unit_price", 0),
                r.get("total_value", 0),
                r.get("location", ""),
                r.get("rider", ""),
                r.get("delivery_charge", 0),
                r.get("payment_method", ""),
                r.get("status", ""),
                str(r.get("created_at", ""))[:16].replace("T", " ")
            ])

        temp_dir = tempfile.gettempdir()
        filepath = os.path.join(temp_dir, "Kenooff_Logistics_Report.xlsx")
        wb.save(filepath)
        return filepath

    def generate_pdf():
        temp_dir = tempfile.gettempdir()
        filepath = os.path.join(temp_dir, "Kenooff_Logistics_Report.pdf")

        doc = SimpleDocTemplate(filepath, pagesize=letter, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle("TitleStyle", parent=styles["Heading1"], fontSize=18, textColor=colors.HexColor("#1F4E78"), spaceAfter=10)
        elements.append(Paragraph("Kenooff Logistics - Package History Report", title_style))
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}", styles["Normal"]))
        elements.append(Spacer(1, 15))

        table_data = [["ID", "Vendor", "Package", "Rider", "Location", "Total (₦)", "Status"]]
        for r in all_rows_cache:
            table_data.append([
                str(r.get("id")),
                str(r.get("vendor", "")),
                str(r.get("package_desc", "")),
                str(r.get("rider", "")),
                str(r.get("location", "")),
                f"₦{r.get('total_value', 0):,.2f}",
                str(r.get("status", ""))
            ])

        t = Table(table_data, colWidths=[30, 80, 110, 80, 90, 80, 70])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
        ]))
        elements.append(t)
        doc.build(elements)
        return filepath

    def open_export_menu(e):
        if not all_rows_cache:
            show_alert("No records available to export!", ft.Colors.RED_600)
            return

        def handle_excel_gen(e):
            try:
                generate_excel()
                show_alert("Excel report created successfully!", ft.Colors.GREEN_800)
            except Exception as ex:
                show_alert(f"Excel Export Error: {str(ex)}", ft.Colors.RED_600)
            dialog.open = False
            page.update()

        def handle_pdf_gen(e):
            try:
                generate_pdf()
                show_alert("PDF report generated successfully!", ft.Colors.GREEN_800)
            except Exception as ex:
                show_alert(f"PDF Export Error: {str(ex)}", ft.Colors.RED_600)
            dialog.open = False
            page.update()

        def close_dialog(e):
            dialog.open = False
            page.update()

        dialog = ft.AlertDialog(
            title=ft.Text("Export Report", size=16, weight=ft.FontWeight.BOLD),
            content=ft.Text("Choose your preferred export format:", size=13),
            actions=[
                ft.ElevatedButton("Excel (.xlsx)", icon=ft.Icons.TABLE_CHART, on_click=handle_excel_gen, style=ft.ButtonStyle(color=ft.Colors.GREEN_800)),
                ft.ElevatedButton("PDF Report (.pdf)", icon=ft.Icons.PICTURE_AS_PDF, on_click=handle_pdf_gen, style=ft.ButtonStyle(color=ft.Colors.RED_800)),
                ft.TextButton("Cancel", on_click=close_dialog),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )

        page.overlay.append(dialog)
        dialog.open = True
        page.update()

    export_btn = ft.IconButton(
        icon=ft.Icons.DOWNLOAD, 
        tooltip="Export Data", 
        icon_color=ft.Colors.BLUE_800,
        on_click=open_export_menu
    )

    history_list = ft.ListView(spacing=10, expand=True)

    def update_analytics(rows):
        total_rev = sum(r.get("total_value", 0.0) or 0.0 for r in rows)
        total_count = len(rows)
        delivered_count = sum(1 for r in rows if r.get("status") == "Delivered")
        pending_count = sum(1 for r in rows if r.get("status") in ["Pending", "In Transit"])

        total_rev_text.value = f"₦{total_rev:,.0f}"
        total_count_text.value = str(total_count)
        delivered_count_text.value = str(delivered_count)
        pending_count_text.value = str(pending_count)

    def render_filtered_cards(e=None):
        history_list.controls.clear()
        query = search_input.value.lower() if search_input.value else ""

        filtered_rows = []
        for r in all_rows_cache:
            if active_filter_status != "All" and r.get("status") != active_filter_status:
                continue
            
            vendor = str(r.get("vendor", "")).lower()
            desc = str(r.get("package_desc", "")).lower()
            rider = str(r.get("rider", "")).lower()
            loc = str(r.get("location", "")).lower()

            if query and query not in vendor and query not in desc and query not in rider and query not in loc:
                continue

            filtered_rows.append(r)

        if not filtered_rows:
            history_list.controls.append(
                ft.Container(
                    content=ft.Text("No matching packages found.", italic=True, color=ft.Colors.GREY_500),
                    alignment=ft.Alignment.CENTER,
                    padding=20
                )
            )
        else:
            for row in filtered_rows:
                pkg_id = row.get("id")
                vendor = row.get("vendor", "")
                desc = row.get("package_desc", "")
                total = row.get("total_value", 0.0) or 0.0
                loc = row.get("location", "")
                rider = row.get("rider", "")
                status = str(row.get("status", "Pending"))
                raw_time = row.get("created_at", "")

                formatted_time = "N/A"
                if raw_time:
                    try:
                        clean_time = raw_time.split(".")[0].replace("T", " ")
                        dt = datetime.strptime(clean_time, "%Y-%m-%d %H:%M:%S")
                        formatted_time = dt.strftime("%b %d, %I:%M %p")
                    except Exception:
                        formatted_time = raw_time[:16].replace("T", " ")

                status_color = ft.Colors.ORANGE_700
                if status == "Delivered":
                    status_color = ft.Colors.GREEN_700
                elif status == "In Transit":
                    status_color = ft.Colors.BLUE_700

                status_menu = ft.PopupMenuButton(
                    content=ft.Container(
                        content=ft.Row([
                            ft.Text(status, color=ft.Colors.BLACK, size=13, weight=ft.FontWeight.W_500),
                            ft.Icon(ft.Icons.ARROW_DROP_DOWN, color=ft.Colors.BLACK54, size=18)
                        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, tight=True),
                        border=ft.Border.all(1, ft.Colors.GREY_400),
                        border_radius=6,
                        padding=ft.Padding(8, 6, 8, 6)
                    ),
                    items=[
                        ft.PopupMenuItem(
                            content=ft.Text("Pending"), 
                            on_click=lambda e, r_id=pkg_id: change_status(r_id, "Pending")
                        ),
                        ft.PopupMenuItem(
                            content=ft.Text("In Transit"), 
                            on_click=lambda e, r_id=pkg_id: change_status(r_id, "In Transit")
                        ),
                        ft.PopupMenuItem(
                            content=ft.Text("Delivered"), 
                            on_click=lambda e, r_id=pkg_id: change_status(r_id, "Delivered")
                        ),
                    ]
                )

                card = ft.Card(
                    content=ft.Container(
                        padding=12,
                        content=ft.Column([
                            ft.Row([
                                ft.Column([
                                    ft.Text(f"{vendor}", weight=ft.FontWeight.BOLD, size=16),
                                    ft.Text(f"Logged: {formatted_time}", size=10, color=ft.Colors.GREY_500)
                                ], spacing=1),
                                ft.Container(
                                    content=ft.Text(f" {status} ", color=ft.Colors.WHITE, size=12),
                                    bgcolor=status_color,
                                    border_radius=4,
                                    padding=4
                                )
                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                            ft.Text(f"Item: {desc}", size=14),
                            ft.Text(f"Location: {loc or 'N/A'} | Rider: {rider or 'N/A'}", size=12, color=ft.Colors.GREY_700),
                            ft.Row([
                                ft.Text(f"₦{total:,.2f}", weight=ft.FontWeight.BOLD, color=ft.Colors.GREEN_800),
                                ft.Row([
                                    status_menu,
                                    ft.IconButton(
                                        icon=ft.Icons.DELETE_OUTLINED, 
                                        icon_color=ft.Colors.RED_400,
                                        on_click=lambda e, r_id=pkg_id: delete_record(r_id)
                                    )
                                ])
                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN)
                        ])
                    )
                )
                history_list.controls.append(card)

        page.update()

    search_input.on_change = render_filtered_cards

    def set_filter(status):
        nonlocal active_filter_status
        active_filter_status = status
        render_filtered_cards()

    filter_buttons = ft.Row([
        ft.OutlinedButton("All", on_click=lambda e: set_filter("All")),
        ft.OutlinedButton("Pending", on_click=lambda e: set_filter("Pending")),
        ft.OutlinedButton("In Transit", on_click=lambda e: set_filter("In Transit")),
        ft.OutlinedButton("Delivered", on_click=lambda e: set_filter("Delivered")),
    ], scroll=ft.ScrollMode.AUTO, spacing=6)

    def delete_record(record_id):
        try:
            supabase_request(f"packages?id=eq.{record_id}", method="DELETE")
            show_alert("Record deleted from Cloud!", ft.Colors.GREY_700)
            load_database_records()
        except Exception as ex:
            show_alert(f"Failed to delete: {str(ex)}", ft.Colors.RED_600)

    def change_status(record_id, new_status):
        try:
            supabase_request(f"packages?id=eq.{record_id}", method="PATCH", data={"status": new_status})
            show_alert(f"Status changed to '{new_status}'!", ft.Colors.GREEN_700)
            load_database_records()
        except Exception as ex:
            show_alert(f"Failed to update status: {str(ex)}", ft.Colors.RED_600)

    def load_database_records():
        nonlocal all_rows_cache
        try:
            all_rows_cache = supabase_request("packages?select=*&order=id.desc", method="GET")
            update_analytics(all_rows_cache)
            render_filtered_cards()
        except Exception as ex:
            history_list.controls.clear()
            history_list.controls.append(
                ft.Container(
                    content=ft.Text(f"Error connecting to Cloud Database: {str(ex)}", color=ft.Colors.RED_500),
                    padding=20
                )
            )
            page.update()

    history_view = ft.Column([
        analytics_row,
        ft.Row([search_input, export_btn], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
        filter_buttons,
        ft.Divider(height=5),
        history_list
    ], expand=True, spacing=10)

    body = ft.Container(content=form_view, padding=12, expand=True)

    def on_nav_change(e):
        idx = int(e.control.selected_index)
        if idx == 0:
            body.content = form_view
        else:
            load_database_records()
            body.content = history_view
        page.update()

    page.navigation_bar = ft.NavigationBar(
        selected_index=0,
        destinations=[
            ft.NavigationBarDestination(icon=ft.Icons.ADD_BOX, label="New Log"),
            ft.NavigationBarDestination(icon=ft.Icons.HISTORY, label="View History"),
        ],
        on_change=on_nav_change
    )

    page.clean()
    page.add(body)
    page.update()

ft.app(target=main)
